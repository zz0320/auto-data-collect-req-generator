from __future__ import annotations

import json
import os
import re
import base64
import hashlib
import hmac
import secrets
import threading
import time
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib import request
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:  # pragma: no cover - shown in /api/health
    openpyxl = None
    Workbook = None
    Alignment = Font = PatternFill = None
    OPENPYXL_IMPORT_ERROR = exc
else:
    OPENPYXL_IMPORT_ERROR = None


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
OUTPUT_DIR = ROOT / "outputs" / "generated"
USER_OUTPUT_ROOT = ROOT / "outputs" / "users"
WORKBOOK_UPLOAD_DIR = ROOT / "work" / "rag_uploads"
USER_DATA_DIR = ROOT / "work" / "users"
AUTH_STORE_PATH = USER_DATA_DIR / "accounts.json"
SESSION_COOKIE_NAME = "rds_session"
SESSION_TTL_SECONDS = 60 * 60 * 12
PASSWORD_HASH_ITERATIONS = 180_000


def load_local_env() -> None:
    env_path = ROOT / ".env.local"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_local_env()

DEFAULT_QWEN_MODEL = os.getenv("QWEN_MODEL", "qwen3.7-max").strip() or "qwen3.7-max"
DEFAULT_QWEN_ENDPOINT = (
    os.getenv("DASHSCOPE_ENDPOINT", "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation").strip()
    or "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation"
)
DEFAULT_SOURCE_XLSX = Path(
    os.getenv(
        "SOURCE_XLSX",
        "/Users/kenton/Downloads/模型训练记录表_数据需求总表同步_收集结果-有效.xlsx",
    )
)


AVATAR_PRESETS = [
    {"id": "young_man", "label": "男生", "row": 0, "col": 0},
    {"id": "young_woman", "label": "女生", "row": 0, "col": 1},
    {"id": "engineer_boy", "label": "工程师男生", "row": 0, "col": 2},
    {"id": "engineer_girl", "label": "工程师女生", "row": 0, "col": 3},
    {"id": "cat", "label": "猫咪", "row": 1, "col": 0},
    {"id": "dog", "label": "狗狗", "row": 1, "col": 1},
    {"id": "rabbit", "label": "兔子", "row": 1, "col": 2},
    {"id": "panda", "label": "熊猫", "row": 1, "col": 3},
    {"id": "robot", "label": "机器人", "row": 2, "col": 0},
    {"id": "fox", "label": "狐狸", "row": 2, "col": 1},
    {"id": "bear", "label": "小熊", "row": 2, "col": 2},
    {"id": "blob", "label": "圆形角色", "row": 2, "col": 3},
]
AVATAR_IDS = {item["id"] for item in AVATAR_PRESETS}
DEFAULT_AVATAR_ID = "robot"


def normalize_avatar_id(value: Any) -> str:
    text = str(value or "").strip()
    return text if text in AVATAR_IDS else DEFAULT_AVATAR_ID


def mask_api_key(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if len(text) <= 10:
        return f"{text[:2]}***{text[-2:]}"
    return f"{text[:3]}***{text[-4:]}"


CURRENT_SOURCE_XLSX = DEFAULT_SOURCE_XLSX


def safe_identifier(value: Any, fallback: str = "user") -> str:
    text = re.sub(r"[^\w.-]+", "_", str(value or "").strip()).strip("._")
    return text[:80] or fallback


def utc_timestamp() -> int:
    return int(time.time())


def hash_password(password: str, salt: str | None = None) -> str:
    if not str(password or ""):
        raise ValueError("密码不能为空")
    salt_text = salt or base64.urlsafe_b64encode(secrets.token_bytes(18)).decode("ascii").rstrip("=")
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        str(password).encode("utf-8"),
        salt_text.encode("utf-8"),
        PASSWORD_HASH_ITERATIONS,
    )
    encoded = base64.urlsafe_b64encode(digest).decode("ascii").rstrip("=")
    return f"pbkdf2_sha256${PASSWORD_HASH_ITERATIONS}${salt_text}${encoded}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations_text, salt, expected = str(stored_hash or "").split("$", 3)
        iterations = int(iterations_text)
    except (ValueError, TypeError):
        return False
    if algorithm != "pbkdf2_sha256" or iterations < 1:
        return False
    digest = hashlib.pbkdf2_hmac("sha256", str(password or "").encode("utf-8"), salt.encode("utf-8"), iterations)
    actual = base64.urlsafe_b64encode(digest).decode("ascii").rstrip("=")
    return hmac.compare_digest(actual, expected)


class AuthStore:
    def __init__(self, path: Path):
        self.path = Path(path)
        self._lock = threading.RLock()

    def _read(self) -> dict[str, Any]:
        if not self.path.exists():
            return {"users": []}
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {"users": []}
        users = data.get("users")
        return {"users": users if isinstance(users, list) else []}

    def _write(self, data: dict[str, Any]) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = self.path.with_suffix(".tmp")
        tmp_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp_path.replace(self.path)

    def _public_user(self, user: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": str(user.get("id") or ""),
            "username": str(user.get("username") or ""),
            "role": str(user.get("role") or "user"),
            "disabled": bool(user.get("disabled", False)),
            "createdAt": user.get("createdAt") or "",
        }

    def _find_user(self, data: dict[str, Any], username: str) -> dict[str, Any] | None:
        wanted = str(username or "").strip().lower()
        for user in data.get("users", []):
            if str(user.get("username") or "").strip().lower() == wanted:
                return user
        return None

    def _find_user_by_id(self, data: dict[str, Any], user_id: str) -> dict[str, Any] | None:
        wanted = str(user_id or "")
        for user in data.get("users", []):
            if str(user.get("id") or "") == wanted:
                return user
        return None

    def bootstrap_admin(self, username: str, password: str) -> dict[str, Any]:
        with self._lock:
            data = self._read()
            for user in data.get("users", []):
                if user.get("role") == "admin":
                    return self._public_user(user)
            username = str(username or "").strip() or "admin"
            user = {
                "id": f"u_{secrets.token_hex(8)}",
                "username": username,
                "role": "admin",
                "passwordHash": hash_password(password),
                "disabled": False,
                "createdAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            data.setdefault("users", []).append(user)
            self._write(data)
            return self._public_user(user)

    def require_admin(self, actor: dict[str, Any]) -> None:
        if str(actor.get("role") or "") != "admin" or actor.get("disabled"):
            raise PermissionError("需要管理员权限")

    def create_user(self, actor: dict[str, Any], username: str, password: str, role: str = "user") -> dict[str, Any]:
        self.require_admin(actor)
        username = str(username or "").strip()
        role = str(role or "user").strip() or "user"
        if not username:
            raise ValueError("用户名不能为空")
        if role not in {"user", "admin"}:
            raise ValueError("角色只能是 user 或 admin")
        with self._lock:
            data = self._read()
            if self._find_user(data, username):
                raise ValueError("用户名已存在")
            user = {
                "id": f"u_{secrets.token_hex(8)}",
                "username": username,
                "role": role,
                "passwordHash": hash_password(password),
                "disabled": False,
                "createdAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            data.setdefault("users", []).append(user)
            self._write(data)
            return self._public_user(user)

    def list_users(self, actor: dict[str, Any]) -> list[dict[str, Any]]:
        self.require_admin(actor)
        with self._lock:
            return [self._public_user(user) for user in self._read().get("users", [])]

    def authenticate(self, username: str, password: str) -> dict[str, Any]:
        with self._lock:
            data = self._read()
            user = self._find_user(data, username)
            if not user or user.get("disabled"):
                raise PermissionError("用户名或密码错误")
            if not verify_password(password, str(user.get("passwordHash") or "")):
                raise PermissionError("用户名或密码错误")
            return self._public_user(user)

    def get_user(self, user_id: str) -> dict[str, Any] | None:
        with self._lock:
            user = self._find_user_by_id(self._read(), user_id)
            if not user or user.get("disabled"):
                return None
            return self._public_user(user)

    def set_user_disabled(self, actor: dict[str, Any], user_id: str, disabled: bool) -> dict[str, Any]:
        self.require_admin(actor)
        with self._lock:
            data = self._read()
            user = self._find_user_by_id(data, user_id)
            if not user:
                raise ValueError("用户不存在")
            if user.get("id") == actor.get("id") and disabled:
                raise ValueError("不能禁用当前管理员账户")
            user["disabled"] = bool(disabled)
            self._write(data)
            return self._public_user(user)

    def set_password(self, actor: dict[str, Any], user_id: str, password: str) -> dict[str, Any]:
        self.require_admin(actor)
        with self._lock:
            data = self._read()
            user = self._find_user_by_id(data, user_id)
            if not user:
                raise ValueError("用户不存在")
            user["passwordHash"] = hash_password(password)
            self._write(data)
            return self._public_user(user)


@dataclass
class WorkbookState:
    source: Path
    summary: dict[str, Any]
    rag_documents: list[dict[str, Any]]


class UserWorkspaceManager:
    def __init__(self, base_dir: Path, default_source: Path, output_root: Path | None = None):
        self.base_dir = Path(base_dir)
        self.default_source = Path(default_source)
        self.output_root = Path(output_root or USER_OUTPUT_ROOT)
        self._lock = threading.RLock()
        self._states: dict[str, WorkbookState] = {}

    def user_id(self, user: dict[str, Any] | None) -> str:
        return safe_identifier((user or {}).get("id") or "local", "local")

    def user_root(self, user: dict[str, Any] | None) -> Path:
        return self.base_dir / self.user_id(user)

    def settings_path(self, user: dict[str, Any] | None) -> Path:
        return self.user_root(user) / "settings.json"

    def _read_settings(self, user: dict[str, Any] | None) -> dict[str, Any]:
        settings = self.settings_path(user)
        if settings.exists():
            try:
                data = json.loads(settings.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                data = {}
            return data if isinstance(data, dict) else {}
        return {}

    def _write_settings(self, user: dict[str, Any] | None, data: dict[str, Any]) -> None:
        root = self.user_root(user)
        root.mkdir(parents=True, exist_ok=True)
        settings = dict(data)
        settings["updatedAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.settings_path(user).write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")

    def _load_source_for_user(self, user: dict[str, Any] | None) -> Path:
        data = self._read_settings(user)
        if data:
            source = data.get("activeSource")
            if source:
                return Path(source)
        return self.default_source

    def _state_from_source(self, source: Path) -> WorkbookState:
        summary = load_workbook_summary(source)
        if not summary.get("ok"):
            raise ValueError(str(summary.get("error") or "存量数据需求表读取失败"))
        docs = summary.pop("ragDocuments", [])
        summary["ragDocumentCount"] = len(docs)
        return WorkbookState(source=Path(source), summary=summary, rag_documents=docs)

    def get_workspace(self, user: dict[str, Any] | None) -> WorkbookState:
        user_id = self.user_id(user)
        with self._lock:
            if user_id not in self._states:
                self._states[user_id] = self._state_from_source(self._load_source_for_user(user))
            return self._states[user_id]

    def set_active_workbook(self, user: dict[str, Any] | None, source_path: Path | str) -> dict[str, Any]:
        source = Path(source_path)
        state = self._state_from_source(source)
        user_id = self.user_id(user)
        with self._lock:
            settings = self._read_settings(user)
            settings["activeSource"] = str(source)
            self._write_settings(user, settings)
            self._states[user_id] = state
        return state.summary

    def qwen_config(self, user: dict[str, Any] | None, include_secret: bool = False) -> dict[str, Any]:
        settings = self._read_settings(user)
        saved = settings.get("qwen") if isinstance(settings.get("qwen"), dict) else {}
        saved_key = str(saved.get("apiKey") or "").strip()
        env_key = os.getenv("DASHSCOPE_API_KEY", "").strip()
        effective_key = saved_key or env_key
        source = "user" if saved_key else ("env" if env_key else "missing")
        model = str(saved.get("model") or DEFAULT_QWEN_MODEL).strip() or DEFAULT_QWEN_MODEL
        endpoint = str(saved.get("endpoint") or DEFAULT_QWEN_ENDPOINT).strip() or DEFAULT_QWEN_ENDPOINT
        config = {
            "ok": True,
            "configured": bool(effective_key),
            "source": source,
            "hasSavedApiKey": bool(saved_key),
            "apiKeyMask": mask_api_key(effective_key),
            "model": model,
            "modelOptions": QWEN_MODEL_OPTIONS,
            "endpoint": endpoint,
            "updatedAt": str(saved.get("updatedAt") or ""),
        }
        if include_secret:
            config["apiKey"] = effective_key
        return config

    def save_qwen_config(
        self,
        user: dict[str, Any] | None,
        api_key: str | None = None,
        model: str | None = None,
        endpoint: str | None = None,
        clear_api_key: bool = False,
    ) -> dict[str, Any]:
        with self._lock:
            settings = self._read_settings(user)
            qwen = settings.get("qwen") if isinstance(settings.get("qwen"), dict) else {}
            qwen = dict(qwen)
            if clear_api_key:
                qwen.pop("apiKey", None)
            elif api_key is not None and str(api_key).strip():
                qwen["apiKey"] = str(api_key).strip()
            clean_model = str(model or "").strip()
            clean_endpoint = str(endpoint or "").strip()
            if clean_model:
                qwen["model"] = clean_model
            if clean_endpoint:
                qwen["endpoint"] = clean_endpoint
            qwen["updatedAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            settings["qwen"] = qwen
            self._write_settings(user, settings)
        return self.qwen_config(user)

    def user_profile(self, user: dict[str, Any] | None) -> dict[str, Any]:
        settings = self._read_settings(user)
        saved = settings.get("profile") if isinstance(settings.get("profile"), dict) else {}
        return {
            "ok": True,
            "avatar": normalize_avatar_id(saved.get("avatar")),
            "avatarOptions": AVATAR_PRESETS,
            "updatedAt": str(saved.get("updatedAt") or ""),
        }

    def save_user_profile(self, user: dict[str, Any] | None, avatar: str | None = None) -> dict[str, Any]:
        with self._lock:
            settings = self._read_settings(user)
            profile = settings.get("profile") if isinstance(settings.get("profile"), dict) else {}
            profile = dict(profile)
            if avatar is not None:
                profile["avatar"] = normalize_avatar_id(avatar)
            profile["updatedAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            settings["profile"] = profile
            self._write_settings(user, settings)
        return self.user_profile(user)

    def save_upload(self, user: dict[str, Any], filename: str, file_data: bytes) -> Path:
        uploads = self.user_root(user) / "rag_uploads"
        uploads.mkdir(parents=True, exist_ok=True)
        saved_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_upload_filename(filename)}"
        saved_path = uploads / saved_name
        saved_path.write_bytes(file_data)
        return saved_path

    def output_dir(self, user: dict[str, Any] | None) -> Path:
        return self.output_root / self.user_id(user) / "generated"

    def resolve_download(self, user: dict[str, Any] | None, filename: str) -> Path:
        safe_name = Path(str(filename or "")).name
        if not safe_name or safe_name != str(filename or ""):
            raise PermissionError("没有访问该文件权限")
        candidate = (self.output_dir(user) / safe_name).resolve()
        root = self.output_dir(user).resolve()
        if root not in candidate.parents or not candidate.exists() or not candidate.is_file():
            raise PermissionError("没有访问该文件权限")
        return candidate


def env_int(name: str, default: int) -> int:
    try:
        value = int(float(os.getenv(name, str(default))))
    except (TypeError, ValueError):
        return default
    return max(value, 1)


QWEN_GENERATION_TIMEOUT_SECONDS = env_int("QWEN_GENERATION_TIMEOUT_SECONDS", 180)
QWEN_IDEA_TIMEOUT_SECONDS = env_int("QWEN_IDEA_TIMEOUT_SECONDS", 180)
QWEN_TEST_TIMEOUT_SECONDS = env_int("QWEN_TEST_TIMEOUT_SECONDS", 30)

QWEN_MODEL_OPTIONS = [
    {
        "id": "qwen3.7-max",
        "label": "Qwen3.7-Max",
        "tag": "最新旗舰",
        "description": "官方模型页当前千问文本生成首位，适合高质量数据需求生成。",
    },
    {
        "id": "qwen3-max",
        "label": "Qwen3-Max",
        "tag": "兼容",
        "description": "当前 DashScope 文本生成接口已验证可用的强模型。",
    },
    {
        "id": "qwen-plus",
        "label": "Qwen-Plus",
        "tag": "稳妥",
        "description": "当前 DashScope 文本生成接口已验证可用，成本和效果均衡。",
    },
    {
        "id": "qwen-turbo",
        "label": "Qwen-Turbo",
        "tag": "低成本",
        "description": "当前 DashScope 文本生成接口已验证可用，适合快速试跑和低成本草稿。",
    },
]

TASK_HEADERS = [
    "自动编号",
    "任务ID",
    "采集时长（小时）",
    "提交时间",
    "提交人",
    "填写日期",
    "任务名称",
    "任务简述",
    "采集设备",
    "采集模式",
    "场景域分类",
    "任务步骤描述",
    "目标次数",
    "数采负责人",
    "机器及环境参数",
    "任务级别",
    "任务步骤数量",
]

CANONICAL_ACTIONS = {
    "Grasp": "Grasp（抓取）",
    "Pick": "Pick（拿起）",
    "Place": "Place（放置）",
    "Release": "Release（释放）",
    "Transfer": "Transfer（转移）",
    "Alignment": "Alignment（对准）",
    "Move": "Move（移动）",
    "Navigate": "Navigate（导航）",
    "Carry": "Carry（携带）",
    "Transport": "Transport（搬运）",
    "Open": "Open（打开）",
    "Close": "Close（关闭）",
    "Pull": "Pull（拉）",
    "Push": "Push（推）",
    "Press": "Press（按压）",
    "Lift": "Lift（抬起）",
    "Lower": "Lower（放下）",
    "Fold": "Fold（折叠）",
    "Unfold": "Unfold（展开）",
    "Straighten": "Straighten（整理）",
    "Flip": "Flip（翻转）",
    "HandOver": "HandOver（传递）",
    "Hold": "Hold（握住）",
    "Pour": "Pour（倒）",
    "Scoop": "Scoop（舀）",
    "Insert": "Insert（插入）",
    "Rotate": "Rotate（旋转）",
    "Pack": "Pack（打包）",
    "Stack": "Stack（堆叠）",
    "Wipe": "Wipe（擦拭）",
    "Touch": "Touch（触摸）",
    "Several Times": "Several Times（多次重复抓取放置）",
    "Crouch": "Crouch（蹲下）",
    "Stretch": "Stretch（伸展）",
    "Plug": "Plug（插插头）",
    "Unplug": "Unplug（拔插头）",
    "Screw": "Screw（拧紧）",
    "Unscrew": "Unscrew（拧松）",
    "Zip": "Zip（拉上）",
    "Unzip": "Unzip（拉开）",
}

KNOWN_CATEGORIES = ["家居家政", "通用抓取放置", "商超药店", "餐饮服务", "工业制造"]
KNOWN_LEVELS = ["简易", "中等", "复杂"]
TASK_PHASES = {
    "pretrain": {
        "label": "预训练",
        "maxTargetTimes": 60,
        "style": "基础能力覆盖，优先生成短步骤、单技能或低组合度任务，覆盖抓取、放置、对准、释放、简单转移等底层能力。",
    },
    "posttrain": {
        "label": "后训练",
        "maxTargetTimes": 600,
        "style": "指令跟随和场景泛化，允许生成更长步骤、多对象、多约束和多场景任务，但仍必须受机器人真实能力限制。",
    },
}
MAX_QWEN_TASKS_PER_BATCH = 30
MAX_GENERATED_TASKS_PER_REQUEST = 200
MOBILE_ACTIONS = {"Move", "Navigate", "Carry", "Transport"}
BIMANUAL_ACTIONS = {"Fold", "Unfold", "HandOver"}
WHOLE_BODY_ACTIONS = {"Crouch", "Stretch"}
FINE_ACTIONS = {"Screw", "Unscrew", "Zip", "Unzip", "Plug", "Unplug", "Twist", "Insert"}
SUCTION_FORBIDDEN_ACTIONS = BIMANUAL_ACTIONS | FINE_ACTIONS | {
    "Pull",
    "Scoop",
    "Pour",
    "Wipe",
    "Hold",
}

BRAND_PREFIXES = [
    "傅利叶",
    "星尘智能",
    "星海图",
    "智元",
    "乐聚",
    "松灵",
    "方舟无限",
    "Franka",
    "UR",
]


def split_device_name(device: str) -> tuple[str, str]:
    text = str(device or "").strip()
    if not text:
        return "", ""
    for prefix in BRAND_PREFIXES:
        if text.startswith(prefix):
            return prefix, text[len(prefix) :].strip()
    match = re.match(r"^([A-Za-z\u4e00-\u9fff]+?)([A-Za-z0-9_-]+)$", text)
    if match and match.group(1) != text:
        return match.group(1), match.group(2)
    return text, ""


def safe_upload_filename(filename: str) -> str:
    name = Path(str(filename or "")).name
    name = re.sub(r"[^\w.\-\u4e00-\u9fff]+", "_", name).strip("._")
    if not name.lower().endswith(".xlsx"):
        raise ValueError("请选择 .xlsx 存量数据表")
    return name or "rag-source.xlsx"


def extract_multipart_file(body: bytes, content_type: str) -> tuple[str, bytes]:
    match = re.search(r"boundary=([^;]+)", content_type or "")
    if not match:
        raise ValueError("上传请求缺少 multipart boundary")
    boundary = match.group(1).strip().strip('"').encode("utf-8")
    marker = b"--" + boundary
    for part in body.split(marker):
        part = part.strip()
        if not part or part == b"--":
            continue
        if part.endswith(b"--"):
            part = part[:-2].rstrip()
        if b"\r\n\r\n" not in part:
            continue
        header_raw, file_data = part.split(b"\r\n\r\n", 1)
        headers = header_raw.decode("utf-8", errors="replace")
        filename_match = re.search(r'filename="([^"]+)"', headers)
        if "name=\"workbook\"" in headers and filename_match:
            return safe_upload_filename(filename_match.group(1)), file_data.rstrip(b"\r\n")
    raise ValueError("上传请求缺少 workbook 文件字段")


def action_key_from_workbook_label(label: str) -> str | None:
    normalized = str(label or "").strip()
    if re.fullmatch(r"\d+(?:\.\d+)?\s*s", normalized, flags=re.IGNORECASE):
        return None
    first = re.split(r"[（(]", normalized, maxsplit=1)[0].strip()
    return first or None


def truncate_text(value: Any, limit: int) -> str:
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def rag_tokenize(*values: Any) -> set[str]:
    text = " ".join(str(value or "").lower() for value in values)
    tokens = set(re.findall(r"[a-z0-9][a-z0-9_-]{1,}", text))
    han_runs = re.findall(r"[\u4e00-\u9fff]{2,}", text)
    for run in han_runs:
        tokens.add(run)
        for size in (2, 3):
            for index in range(max(len(run) - size + 1, 0)):
                tokens.add(run[index : index + size])
    return tokens


def build_rag_document(item: dict[str, Any], serial: int) -> dict[str, Any]:
    steps = str(item.get("任务步骤描述") or "")
    actions: list[str] = []
    for action_label in re.findall(r"<([^<>]+)>", steps):
        action_key = action_key_from_workbook_label(action_label)
        if action_key and action_key in CANONICAL_ACTIONS and action_key not in actions:
            actions.append(action_key)
    doc = {
        "serial": serial,
        "任务名称": str(item.get("任务名称") or "").strip(),
        "任务简述": str(item.get("任务简述") or "").strip(),
        "采集设备": str(item.get("采集设备") or "").strip(),
        "采集模式": str(item.get("采集模式") or "").strip(),
        "场景域分类": str(item.get("场景域分类") or "").strip(),
        "任务步骤描述": steps,
        "目标次数": item.get("目标次数") or "",
        "任务级别": str(item.get("任务级别") or "").strip(),
        "任务步骤数量": item.get("任务步骤数量") or "",
        "动作标签": actions,
    }
    doc["_tokens"] = rag_tokenize(
        doc["任务名称"],
        doc["任务简述"],
        doc["采集设备"],
        doc["采集模式"],
        doc["场景域分类"],
        doc["任务步骤描述"],
        " ".join(actions),
        " ".join(CANONICAL_ACTIONS.get(action, action) for action in actions),
    )
    return doc


def compact_rag_example(doc: dict[str, Any], idea: str = "", max_steps: int = 320) -> dict[str, Any]:
    return {
        "匹配idea": idea,
        "任务名称": doc.get("任务名称", ""),
        "任务简述": truncate_text(doc.get("任务简述", ""), 120),
        "采集设备": doc.get("采集设备", ""),
        "采集模式": doc.get("采集模式", ""),
        "场景域分类": doc.get("场景域分类", ""),
        "任务级别": doc.get("任务级别", ""),
        "目标次数": doc.get("目标次数", ""),
        "任务步骤数量": doc.get("任务步骤数量", ""),
        "动作标签": doc.get("动作标签", []),
        "任务步骤描述": truncate_text(doc.get("任务步骤描述", ""), max_steps),
    }


def retrieve_rag_examples(
    idea: str,
    robots: list["RobotProfile"],
    rag_documents: list[dict[str, Any]] | None = None,
    limit: int = 6,
) -> list[dict[str, Any]]:
    documents = rag_documents if rag_documents is not None else RAG_DOCUMENTS
    if not documents or limit <= 0:
        return []

    idea_tokens = rag_tokenize(idea)
    robot_identity_tokens = rag_tokenize(" ".join(f"{robot.name} {robot.brand} {robot.model}" for robot in robots))
    query_tokens = idea_tokens or robot_identity_tokens
    robot_names = {robot.name for robot in robots}
    robot_brands = {robot.brand.strip() for robot in robots if robot.brand.strip()}
    robot_modes = {robot.arms for robot in robots if robot.arms}
    idea_text = str(idea or "").strip()
    scored: list[tuple[float, dict[str, Any]]] = []

    for doc in documents:
        doc_tokens = doc.get("_tokens", set())
        overlap = query_tokens & doc_tokens
        idea_overlap = idea_tokens & doc_tokens
        score = float(len(overlap)) + len(idea_overlap) * 3.0
        doc_text = " ".join(
            str(doc.get(field) or "")
            for field in ["任务名称", "任务简述", "采集设备", "采集模式", "场景域分类", "任务步骤描述"]
        )
        if idea_text and idea_text in doc_text:
            score += 8.0
        device = str(doc.get("采集设备") or "")
        if device in robot_names:
            score += 2.0 if idea_overlap else 0.7
        elif any(brand and brand in device for brand in robot_brands):
            score += 1.0 if idea_overlap else 0.3
        if doc.get("采集模式") in robot_modes:
            score += 0.8 if idea_overlap else 0.2
        if score > 0:
            scored.append((score, doc))

    scored.sort(key=lambda item: (-item[0], str(item[1].get("任务名称") or "")))
    return [compact_rag_example(doc, idea) for _, doc in scored[:limit]]


def build_rag_context(
    robots: list["RobotProfile"],
    ideas: list[str],
    rag_documents: list[dict[str, Any]] | None = None,
    limit_per_idea: int = 4,
    total_limit: int = 12,
) -> dict[str, Any]:
    selected: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    query_ideas = ideas or [robot.name for robot in robots] or ["机器人数据采集"]
    for idea in query_ideas:
        for example in retrieve_rag_examples(idea, robots, rag_documents, limit=limit_per_idea):
            key = (str(example.get("任务名称") or ""), str(example.get("采集设备") or ""))
            if key in seen:
                continue
            selected.append(example)
            seen.add(key)
            if len(selected) >= total_limit:
                break
        if len(selected) >= total_limit:
            break
    return {
        "enabled": bool(selected),
        "indexSize": len(rag_documents if rag_documents is not None else RAG_DOCUMENTS),
        "retrievalMethod": "local_keyword_bm25_like",
        "examples": selected,
    }


@dataclass
class RobotProfile:
    brand: str
    model: str
    end_effector: str
    arms: str
    mobile: bool
    whole_body: bool
    notes: str

    @property
    def name(self) -> str:
        brand = self.brand.strip()
        model = self.model.strip()
        if brand and model:
            return f"{brand}{model}" if model.startswith(brand) else f"{brand}{model}"
        return brand or model or "未命名机器人"

    @property
    def is_dual_arm(self) -> bool:
        return self.arms == "双臂"

    @property
    def is_left_only(self) -> bool:
        return self.arms == "单臂_左"

    @property
    def is_right_only(self) -> bool:
        return self.arms == "单臂_右"

    @property
    def is_dexterous(self) -> bool:
        text = f"{self.end_effector} {self.notes}"
        return any(token in text for token in ["灵巧手", "五指", "多指", "dexterous", "Dexterous"])

    @property
    def has_manipulator(self) -> bool:
        return bool(self.end_effector.strip()) and self.end_effector.strip() not in ["无", "仅相机", "摄像头"]

    def summary(self) -> str:
        mobile = "可移动" if self.mobile else "固定工位"
        body = "具备全身/蹲伸能力" if self.whole_body else "不假设全身能力"
        return f"{self.name}；{self.arms}；末端执行器：{self.end_effector or '未填写'}；{mobile}；{body}；备注：{self.notes or '无'}"


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today_midnight_text() -> str:
    return datetime.now().strftime("%Y-%m-%d 00:00:00")


def load_workbook_summary(source_path: Path | None = None) -> dict[str, Any]:
    source = Path(source_path or CURRENT_SOURCE_XLSX)
    if OPENPYXL_IMPORT_ERROR:
        return {"ok": False, "error": f"openpyxl unavailable: {OPENPYXL_IMPORT_ERROR}"}
    if not source.exists():
        return {"ok": False, "error": f"source workbook not found: {source}"}

    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [header for header in raw_headers if header]
    header_index = {header: idx for idx, header in enumerate(raw_headers) if header}
    devices: dict[str, int] = {}
    categories: dict[str, int] = {}
    modes: dict[str, int] = {}
    levels: dict[str, int] = {}
    robot_stats: dict[str, dict[str, Any]] = {}
    examples: list[dict[str, Any]] = []
    rag_documents: list[dict[str, Any]] = []
    max_auto_id = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        device_idx = header_index.get("采集设备")
        mode_idx = header_index.get("采集模式")
        category_idx = header_index.get("场景域分类")
        steps_idx = header_index.get("任务步骤描述")
        device_value = str(row[device_idx] or "").strip() if device_idx is not None and device_idx < len(row) else ""
        mode_value = str(row[mode_idx] or "").strip() if mode_idx is not None and mode_idx < len(row) else ""
        category_value = str(row[category_idx] or "").strip() if category_idx is not None and category_idx < len(row) else ""
        steps_value = str(row[steps_idx] or "") if steps_idx is not None and steps_idx < len(row) else ""

        if device_value:
            stats = robot_stats.setdefault(
                device_value,
                {
                    "count": 0,
                    "modes": Counter(),
                    "categories": Counter(),
                    "actions": Counter(),
                    "mobile_mentions": 0,
                    "whole_body_mentions": 0,
                },
            )
            stats["count"] += 1
            if mode_value:
                stats["modes"][mode_value] += 1
            if category_value:
                for category_part in re.split(r"[,，]\s*", category_value):
                    if category_part:
                        stats["categories"][category_part] += 1
            for action_label in re.findall(r"<([^<>]+)>", steps_value):
                action_key = action_key_from_workbook_label(action_label)
                if action_key:
                    stats["actions"][action_key] += 1
            if re.search(r"机器人移动到|底盘|导航到|巡检|跨房间|移动到[^。；\n]*(货架|分拣台|桌子旁|目标点)", steps_value):
                stats["mobile_mentions"] += 1
            if re.search(r"蹲下|弯腰|地面|低柜|高柜|脚边", steps_value):
                stats["whole_body_mentions"] += 1

        auto_value = row[header_index.get("自动编号", 0)]
        if isinstance(auto_value, (int, float)):
            max_auto_id = max(max_auto_id, int(auto_value))
        elif isinstance(auto_value, str) and auto_value.strip().isdigit():
            max_auto_id = max(max_auto_id, int(auto_value.strip()))
        for column, bucket in [
            ("采集设备", devices),
            ("场景域分类", categories),
            ("采集模式", modes),
            ("任务级别", levels),
        ]:
            idx = header_index.get(column)
            value = row[idx] if idx is not None and idx < len(row) else None
            if value:
                bucket[str(value)] = bucket.get(str(value), 0) + 1
        item = {
            header: row[header_index[header]]
            for header in TASK_HEADERS
            if header in header_index and header_index[header] < len(row)
        }
        if item.get("任务名称") and item.get("任务步骤描述"):
            rag_documents.append(build_rag_document(item, len(rag_documents) + 1))
            if len(examples) < 8:
                examples.append(item)

    robot_presets = []
    for device, stats in sorted(robot_stats.items(), key=lambda item: item[1]["count"], reverse=True):
        actions = stats["actions"]
        dominant_mode = stats["modes"].most_common(1)[0][0] if stats["modes"] else "双臂"
        common_categories = [name for name, _ in stats["categories"].most_common(3)]
        common_actions = [CANONICAL_ACTIONS.get(name, name) for name, _ in actions.most_common(5)]
        brand, model = split_device_name(device)
        mobile = stats["mobile_mentions"] > 0
        whole_body = stats["whole_body_mentions"] > 0 or any(actions.get(action, 0) > 0 for action in WHOLE_BODY_ACTIONS)
        dexterous_evidence = {"Screw", "Unscrew", "Zip", "Unzip", "Plug", "Unplug"}
        end_effector = "灵巧手" if any(actions.get(action, 0) > 0 for action in dexterous_evidence) else "夹爪"
        notes_parts = [f"从存量 {stats['count']} 条需求归纳"]
        if common_categories:
            notes_parts.append(f"常见场景：{'、'.join(common_categories)}")
        if common_actions:
            notes_parts.append(f"常见动作：{'、'.join(common_actions[:3])}")
        if not mobile:
            notes_parts.append("未从存量动作中观察到移动底盘任务，默认按固定工位处理")
        if not whole_body:
            notes_parts.append("未从存量动作中观察到全身/蹲伸任务，默认不假设全身能力")
        robot_presets.append(
            {
                "name": device,
                "brand": brand,
                "model": model,
                "endEffector": end_effector,
                "arms": dominant_mode if dominant_mode in ["双臂", "单臂_左", "单臂_右"] else "双臂",
                "mobile": mobile,
                "wholeBody": whole_body,
                "notes": "；".join(notes_parts),
                "count": stats["count"],
                "categories": common_categories,
                "actions": common_actions,
                "modes": stats["modes"].most_common(3),
            }
        )

    return {
        "ok": True,
        "source": str(source),
        "sheet": ws.title,
        "rows": max(ws.max_row - 1, 0),
        "columns": ws.max_column,
        "headers": headers[:17],
        "max_auto_id": max_auto_id,
        "devices": sorted(devices.items(), key=lambda x: x[1], reverse=True)[:30],
        "categories": sorted(categories.items(), key=lambda x: x[1], reverse=True),
        "modes": sorted(modes.items(), key=lambda x: x[1], reverse=True),
        "levels": sorted(levels.items(), key=lambda x: x[1], reverse=True),
        "actions": list(CANONICAL_ACTIONS.values()),
        "robotPresets": robot_presets,
        "taskPhases": TASK_PHASES,
        "qwenModelOptions": QWEN_MODEL_OPTIONS,
        "examples": examples,
        "ragDocuments": rag_documents,
    }


def set_active_workbook(source_path: Path | str) -> dict[str, Any]:
    global CURRENT_SOURCE_XLSX, WORKBOOK_SUMMARY, RAG_DOCUMENTS
    source = Path(source_path)
    summary = load_workbook_summary(source)
    if not summary.get("ok"):
        raise ValueError(str(summary.get("error") or "存量数据需求表读取失败"))
    docs = summary.pop("ragDocuments", [])
    summary["ragDocumentCount"] = len(docs)
    CURRENT_SOURCE_XLSX = source
    WORKBOOK_SUMMARY = summary
    RAG_DOCUMENTS = docs
    return WORKBOOK_SUMMARY


WORKBOOK_SUMMARY: dict[str, Any] = {}
RAG_DOCUMENTS: list[dict[str, Any]] = []
try:
    set_active_workbook(DEFAULT_SOURCE_XLSX)
except ValueError as exc:
    WORKBOOK_SUMMARY = {"ok": False, "error": str(exc), "source": str(DEFAULT_SOURCE_XLSX)}
    RAG_DOCUMENTS = []


AUTH_STORE = AuthStore(AUTH_STORE_PATH)
DEFAULT_ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin").strip() or "admin"
DEFAULT_ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123456").strip() or "admin123456"
try:
    BOOTSTRAP_ADMIN = AUTH_STORE.bootstrap_admin(DEFAULT_ADMIN_USERNAME, DEFAULT_ADMIN_PASSWORD)
except ValueError:
    BOOTSTRAP_ADMIN = {"id": "", "username": DEFAULT_ADMIN_USERNAME, "role": "admin", "disabled": False, "createdAt": ""}

WORKSPACE_MANAGER = UserWorkspaceManager(USER_DATA_DIR, DEFAULT_SOURCE_XLSX, USER_OUTPUT_ROOT)
LOCAL_USER = {"id": "local", "username": "local", "role": "admin", "disabled": False, "createdAt": ""}
SESSION_STORE: dict[str, dict[str, Any]] = {}
SESSION_LOCK = threading.RLock()


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "on", "是"}
    return bool(value)


def public_session_user(user: dict[str, Any] | None, workspace_manager: UserWorkspaceManager | None = None) -> dict[str, Any] | None:
    if not user:
        return None
    profile = (workspace_manager or WORKSPACE_MANAGER).user_profile(user) if workspace_manager is not None else {"avatar": DEFAULT_AVATAR_ID}
    return {
        "id": str(user.get("id") or ""),
        "username": str(user.get("username") or ""),
        "role": str(user.get("role") or "user"),
        "isAdmin": str(user.get("role") or "") == "admin",
        "avatar": normalize_avatar_id(profile.get("avatar")),
    }


def create_session(user: dict[str, Any]) -> str:
    token = secrets.token_urlsafe(32)
    with SESSION_LOCK:
        SESSION_STORE[token] = {"userId": user["id"], "expiresAt": utc_timestamp() + SESSION_TTL_SECONDS}
    return token


def destroy_session(token: str) -> None:
    with SESSION_LOCK:
        SESSION_STORE.pop(str(token or ""), None)


def user_from_session_token(token: str) -> dict[str, Any] | None:
    token = str(token or "")
    if not token:
        return None
    with SESSION_LOCK:
        session = SESSION_STORE.get(token)
        if not session:
            return None
        if int(session.get("expiresAt") or 0) <= utc_timestamp():
            SESSION_STORE.pop(token, None)
            return None
        user = AUTH_STORE.get_user(str(session.get("userId") or ""))
        if not user:
            SESSION_STORE.pop(token, None)
            return None
        session["expiresAt"] = utc_timestamp() + SESSION_TTL_SECONDS
        return user


def parse_session_cookie(header_value: str | None) -> str:
    if not header_value:
        return ""
    cookie = SimpleCookie()
    try:
        cookie.load(header_value)
    except Exception:
        return ""
    morsel = cookie.get(SESSION_COOKIE_NAME)
    return morsel.value if morsel else ""


def session_cookie_header(token: str, max_age: int = SESSION_TTL_SECONDS) -> str:
    return f"{SESSION_COOKIE_NAME}={token}; Path=/; Max-Age={max_age}; HttpOnly; SameSite=Lax"


def clear_session_cookie_header() -> str:
    return f"{SESSION_COOKIE_NAME}=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax"


def workspace_state_for(
    current_user: dict[str, Any] | None = None,
    workspace_manager: UserWorkspaceManager | None = None,
) -> WorkbookState | None:
    if current_user is None and workspace_manager is None:
        return None
    return (workspace_manager or WORKSPACE_MANAGER).get_workspace(current_user or LOCAL_USER)


def qwen_config_for_user(
    current_user: dict[str, Any] | None = None,
    workspace_manager: UserWorkspaceManager | None = None,
    include_secret: bool = False,
) -> dict[str, Any]:
    if current_user is None and workspace_manager is None:
        env_key = os.getenv("DASHSCOPE_API_KEY", "").strip()
        config = {
            "ok": True,
            "configured": bool(env_key),
            "source": "env" if env_key else "missing",
            "hasSavedApiKey": False,
            "apiKeyMask": mask_api_key(env_key),
            "model": DEFAULT_QWEN_MODEL,
            "modelOptions": QWEN_MODEL_OPTIONS,
            "endpoint": DEFAULT_QWEN_ENDPOINT,
            "updatedAt": "",
        }
        if include_secret:
            config["apiKey"] = env_key
        return config
    return (workspace_manager or WORKSPACE_MANAGER).qwen_config(current_user or LOCAL_USER, include_secret=include_secret)


def resolve_qwen_request_config(
    body: dict[str, Any],
    current_user: dict[str, Any] | None = None,
    workspace_manager: UserWorkspaceManager | None = None,
    allow_body_api_key: bool = False,
) -> tuple[str, str, str, str]:
    config = qwen_config_for_user(current_user, workspace_manager, include_secret=True)
    body_key = str(body.get("qwenApiKey") or body.get("apiKey") or "").strip() if allow_body_api_key else ""
    api_key = body_key or str(config.get("apiKey") or "").strip()
    model = str(body.get("qwenModel") or config.get("model") or DEFAULT_QWEN_MODEL).strip() or DEFAULT_QWEN_MODEL
    endpoint = str(body.get("qwenEndpoint") or config.get("endpoint") or DEFAULT_QWEN_ENDPOINT).strip() or DEFAULT_QWEN_ENDPOINT
    if not api_key:
        raise ValueError("请先在 API 配置中填写 DashScope API Key")
    source = "draft" if body_key else str(config.get("source") or "missing")
    return api_key, model, endpoint, source


def qwen_config_response(
    current_user: dict[str, Any] | None = None,
    workspace_manager: UserWorkspaceManager | None = None,
) -> dict[str, Any]:
    return qwen_config_for_user(current_user, workspace_manager, include_secret=False)


def save_qwen_config_response(
    body: dict[str, Any],
    current_user: dict[str, Any] | None = None,
    workspace_manager: UserWorkspaceManager | None = None,
) -> dict[str, Any]:
    manager = workspace_manager or WORKSPACE_MANAGER
    return manager.save_qwen_config(
        current_user or LOCAL_USER,
        api_key=str(body.get("apiKey") or body.get("qwenApiKey") or "").strip() or None,
        model=str(body.get("model") or body.get("qwenModel") or "").strip() or None,
        endpoint=str(body.get("endpoint") or body.get("qwenEndpoint") or "").strip() or None,
        clear_api_key=parse_bool(body.get("clearApiKey")),
    )


def qwen_config_test_response(
    body: dict[str, Any],
    current_user: dict[str, Any] | None = None,
    workspace_manager: UserWorkspaceManager | None = None,
) -> dict[str, Any]:
    api_key, model, endpoint, source = resolve_qwen_request_config(
        {
            "qwenApiKey": body.get("apiKey") or body.get("qwenApiKey"),
            "qwenModel": body.get("model") or body.get("qwenModel"),
            "qwenEndpoint": body.get("endpoint") or body.get("qwenEndpoint"),
        },
        current_user=current_user,
        workspace_manager=workspace_manager,
        allow_body_api_key=True,
    )
    parsed = call_qwen_json(
        "你是 API 连通性测试器，只输出 JSON。",
        '请只输出 {"ok": true}，不要添加解释。',
        api_key,
        model,
        endpoint,
        timeout=QWEN_TEST_TIMEOUT_SECONDS,
    )
    return {
        "ok": True,
        "configured": True,
        "source": source,
        "model": model,
        "endpoint": endpoint,
        "result": bool(parsed.get("ok", True)) if isinstance(parsed, dict) else True,
    }


def parse_task_phase(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"posttrain", "post", "after", "后训练", "后训练任务"}:
        return "posttrain"
    return "pretrain"


def workbook_context_for_prompt(
    example_limit: int = 8,
    workbook_summary: dict[str, Any] | None = None,
    rag_documents: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    summary = workbook_summary if workbook_summary is not None else WORKBOOK_SUMMARY
    docs = rag_documents if rag_documents is not None else RAG_DOCUMENTS
    examples = []
    for item in summary.get("examples", [])[:example_limit]:
        examples.append(
            {
                "任务名称": item.get("任务名称"),
                "采集设备": item.get("采集设备"),
                "采集模式": item.get("采集模式"),
                "场景域分类": item.get("场景域分类"),
                "任务级别": item.get("任务级别"),
                "任务简述": item.get("任务简述"),
                "任务步骤描述": str(item.get("任务步骤描述") or "")[:260],
            }
        )
    return {
        "rows": summary.get("rows", 0),
        "topDevices": summary.get("devices", [])[:12],
        "topCategories": summary.get("categories", [])[:8],
        "topModes": summary.get("modes", [])[:5],
        "topLevels": summary.get("levels", [])[:5],
        "actions": summary.get("actions", [])[:30],
        "ragDocumentCount": summary.get("ragDocumentCount", len(docs)),
        "examples": examples,
    }


def parse_robots(raw: Any) -> list[RobotProfile]:
    robots = raw if isinstance(raw, list) else []
    parsed: list[RobotProfile] = []
    for item in robots:
        if not isinstance(item, dict):
            continue
        parsed.append(
            RobotProfile(
                brand=str(item.get("brand", "")).strip(),
                model=str(item.get("model", "")).strip(),
                end_effector=str(item.get("endEffector", item.get("end_effector", ""))).strip(),
                arms=str(item.get("arms", "双臂")).strip() or "双臂",
                mobile=parse_bool(item.get("mobile", False)),
                whole_body=parse_bool(item.get("wholeBody", item.get("whole_body", False))),
                notes=str(item.get("notes", "")).strip(),
            )
        )
    return parsed


def split_task_ideas(value: Any) -> list[str]:
    if isinstance(value, list):
        ideas = [str(item).strip() for item in value]
    else:
        text = str(value or "")
        ideas = [part.strip(" -\t\r\n") for part in re.split(r"[\n；;]+", text)]
    return [idea for idea in ideas if idea]


def derive_capabilities(robot: RobotProfile) -> dict[str, Any]:
    allowed = {"Grasp", "Pick", "Place", "Release", "Transfer", "Alignment", "Press", "Push", "Pull", "Open", "Close", "Lift", "Lower", "Flip", "Rotate", "Touch", "Several Times"}
    blocked: list[str] = []
    cautions: list[str] = []

    if not robot.has_manipulator:
        allowed = set()
        blocked.append("未配置可操作末端执行器，不能生成抓取/放置类数据采集任务")

    if robot.mobile:
        allowed.update(MOBILE_ACTIONS)
    else:
        blocked.append("未配置移动能力，禁止生成跨位置移动、货架往返、巡检搬运任务")

    if robot.is_dual_arm:
        allowed.update({"HandOver", "Hold", "Fold", "Unfold", "Straighten", "Pack", "Stack", "Wipe"})
    else:
        blocked.append("单臂配置不能生成需要双手协作的折叠、双手传递、双手保持任务")

    if robot.whole_body:
        allowed.update(WHOLE_BODY_ACTIONS)
    else:
        cautions.append("未配置全身能力时，只生成桌面或工位高度任务")

    if robot.is_dexterous:
        allowed.update(FINE_ACTIONS | {"Scoop", "Pour", "Wipe"})
    elif "吸盘" in robot.end_effector:
        allowed.difference_update(SUCTION_FORBIDDEN_ACTIONS)
        cautions.append("吸盘优先生成扁平、硬质、表面可吸附物体任务")
    else:
        cautions.append("非灵巧手配置不生成拧螺丝、拉拉链、插拔插头等精细任务")

    return {
        "name": robot.name,
        "summary": robot.summary(),
        "allowedActions": sorted(CANONICAL_ACTIONS[action] for action in allowed if action in CANONICAL_ACTIONS),
        "blocked": blocked,
        "cautions": cautions,
    }


def infer_category(idea: str) -> str:
    if re.search(r"药|商品|货架|超市|商店|扫码|补货", idea):
        return "商超药店"
    if re.search(r"餐|碗|盘|杯|筷|勺|厨房|面包|饮料|咖啡", idea):
        return "餐饮服务"
    if re.search(r"电池|工件|装配|螺丝|线束|产线|包装盒|质检", idea):
        return "工业制造"
    if re.search(r"抓取|放置|摆放|分拣", idea):
        return "通用抓取放置"
    return "家居家政"


def choose_mode(robot: RobotProfile, idea: str) -> str:
    if robot.is_dual_arm and re.search(r"折|叠|双手|配对|铺平|传递|大件", idea):
        return "双臂"
    if robot.is_left_only:
        return "单臂_左"
    if robot.is_right_only:
        return "单臂_右"
    return "双臂" if robot.is_dual_arm else "单臂_右"


def task_level(step_count: int, actions: list[str]) -> str:
    action_set = set(actions)
    if step_count >= 8 or action_set & (BIMANUAL_ACTIONS | WHOLE_BODY_ACTIONS | FINE_ACTIONS):
        return "复杂"
    if step_count >= 5 or action_set & {"Open", "Close", "Pour", "Scoop", "Wipe", "Navigate", "Move"}:
        return "中等"
    return "简易"


def target_for_level(level: str, max_target_times: int | None = None) -> int:
    target = {"简易": 300, "中等": 180, "复杂": 80}.get(level, 120)
    if max_target_times is not None:
        target = min(target, max_target_times)
    return max(target, 1)


def clean_task_name(text: str) -> str:
    text = re.sub(r"\s+", "", text.strip())
    text = re.sub(r"[^\w\u4e00-\u9fff_-]+", "", text)
    return text[:24] or "任务"


def task_name_phase_prefix(task_phase: str) -> str:
    return "后-" if parse_task_phase(task_phase) == "posttrain" else "预-"


def normalize_task_name_phase(value: Any, task_phase: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^(?:[【\[]?(?:预训练|后训练)[】\]]?\s*[:：、-]?\s*)+", "", text)
    text = re.sub(r"^(?:预-|后-)+", "", text).strip()
    return f"{task_name_phase_prefix(task_phase)}{text}" if text else text


def strip_phase_from_brief(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^(?:[【\[]?(?:预训练|后训练)[】\]]?\s*[:：、-]?\s*)+", "", text)
    text = text.replace("【预训练】", "").replace("【后训练】", "")
    return re.sub(r"\s+", " ", text).strip()


def format_steps(items: list[tuple[str, str, int]]) -> str:
    lines = []
    for index, (description, action, seconds) in enumerate(items, start=1):
        action_text = CANONICAL_ACTIONS.get(action, action)
        lines.append(f"{index}. {description} <{action_text}><{seconds}s>")
    return "\n".join(lines)


def seconds_from_steps(steps: str) -> int:
    total = 0
    for match in re.finditer(r"<\s*(\d+(?:\.\d+)?)\s*s\s*>", steps, flags=re.IGNORECASE):
        total += int(float(match.group(1)))
    return total


def line_count_from_steps(steps: str) -> int:
    lines = [line for line in str(steps or "").splitlines() if line.strip()]
    if lines:
        return len(lines)
    actions = extract_action_keys(steps)
    return len(actions)


def build_task_row(
    robot: RobotProfile,
    idea: str,
    index: int,
    max_auto_id: int,
    steps: list[tuple[str, str, int]],
) -> dict[str, Any]:
    step_text = format_steps(steps)
    actions = [action for _, action, _ in steps]
    level = task_level(len(steps), actions)
    target = target_for_level(level)
    name = f"{clean_task_name(idea)}_{index}"
    workspace = "可移动场景" if robot.mobile else "固定工位"
    brief = f"基于“{idea}”生成，限定在{workspace}和已配置末端执行器能力内完成。"
    return {
        "自动编号": "",
        "任务ID": "",
        "采集时长（小时）": "",
        "提交时间": now_text(),
        "提交人": "AI需求生成器",
        "填写日期": today_midnight_text(),
        "任务名称": name,
        "任务简述": brief,
        "采集设备": robot.name,
        "采集模式": choose_mode(robot, idea),
        "场景域分类": infer_category(idea),
        "任务步骤描述": step_text,
        "目标次数": target,
        "数采负责人": "",
        "机器及环境参数": robot.summary(),
        "任务级别": level,
        "任务步骤数量": len(steps),
    }


def qwen_prompt(
    robots: list[RobotProfile],
    ideas: list[str],
    task_count: int,
    task_phase: str,
    batch_index: int = 1,
    workspace_state: WorkbookState | None = None,
) -> tuple[str, str]:
    phase = TASK_PHASES[task_phase]
    max_target_times = int(phase["maxTargetTimes"])
    schema = {
        "headers": TASK_HEADERS,
        "categories": KNOWN_CATEGORIES,
        "levels": KNOWN_LEVELS,
        "actions": list(CANONICAL_ACTIONS.values()),
        "format": "任务步骤描述必须逐行编号，每一行末尾必须包含 <动作（中文）><秒数s>。",
    }
    workbook_summary = workspace_state.summary if workspace_state else WORKBOOK_SUMMARY
    rag_documents = workspace_state.rag_documents if workspace_state else RAG_DOCUMENTS
    workbook_context = workbook_context_for_prompt(example_limit=6, workbook_summary=workbook_summary, rag_documents=rag_documents)
    rag_context = build_rag_context(robots, ideas, rag_documents=rag_documents, limit_per_idea=4, total_limit=12)
    robot_text = "\n".join(
        f"- {robot.summary()}\n  允许动作：{', '.join(derive_capabilities(robot)['allowedActions'])}"
        for robot in robots
    )
    ideas_text = "\n".join(f"- {idea}" for idea in ideas) or "- 根据存量任务风格生成保守的桌面操作任务"
    system = (
        "你是机器人自动数据采集需求生成器。你必须严守机器人实际能力，不确定就不要生成。"
        "禁止编造移动、全身、双臂、灵巧操作能力。输出只能是 JSON，不要 Markdown。"
    )
    user = f"""
请根据存量数据需求表格式，生成新的数据采集需求。

任务阶段：{phase["label"]}
阶段策略：{phase["style"]}
单条任务目标次数上限：{max_target_times} 次
本批次编号：{batch_index}
本批次必须生成任务需求条数：{task_count}
任务名称阶段前缀：{task_name_phase_prefix(task_phase)}

机器人配置：
{robot_text}

新的任务 idea：
{ideas_text}

存量数据摘要：
{json.dumps(workbook_context, ensure_ascii=False, indent=2)}

存量数据 RAG 检索上下文（优先参考这些相似历史任务的字段写法、步骤粒度、动作标签和目标次数范围，但不要照抄任务名或步骤）：
{json.dumps(rag_context, ensure_ascii=False, indent=2)}

字段和动作要求：
{json.dumps(schema, ensure_ascii=False, indent=2)}

硬性规则：
0. 自动编号、任务ID、采集时长（小时）不需要填写，留空；不要自行生成编号、UUID 或采集时长。
1. 固定工位机器人不能包含 Move（移动）、Navigate（导航）、Carry（携带）、Transport（搬运）等底盘移动任务。
2. 单臂机器人不能生成 Fold（折叠）、Unfold（展开）、HandOver（传递）等明显双臂协作任务。
3. 未配置全身能力时，不生成蹲下、伸展、地面拾取、低柜/高柜任务。
4. 吸盘末端只生成扁平硬质、可吸附物体任务，不生成布料折叠、液体倾倒、拧螺丝、拉链、插拔插头。
5. 非灵巧手不要生成拧螺丝、拉拉链、插拔插头等精细手指操作。
6. 任务步骤数量要和任务步骤描述行数一致；目标次数必须在 1 到 {max_target_times} 之间，这是单条任务要采集的次数，不是生成任务需求的条数。
7. 输出 tasks 数组长度必须等于“本批次必须生成任务需求条数”。
8. 同一批内任务名称不能重复；不同机器人之间要结合各自能力差异，不要机械复制同一任务。
9. 任务简述不要写【预训练】或【后训练】；任务名称必须以“{task_name_phase_prefix(task_phase)}”开头。
10. 优先参考 RAG 样例的采集设备、场景域、动作标签和步骤粒度；禁止简单复制历史任务名称和步骤。
11. 严禁生成与存量需求重复或近似重复的任务；任务名称、采集设备、任务简述和步骤组合不得与 RAG 样例一致。如 idea 与历史任务接近，必须更换对象、场景约束或动作组合后再生成。

输出 JSON schema：
{{
  "tasks": [
    {{
      "任务名称": "{task_name_phase_prefix(task_phase)}string",
      "任务简述": "string，不包含【预训练】或【后训练】",
      "采集设备": "必须等于某个机器人名称",
      "采集模式": "双臂|单臂_右|单臂_左",
      "场景域分类": "家居家政|通用抓取放置|商超药店|餐饮服务|工业制造",
      "任务步骤描述": "1. ... <Grasp（抓取）><8s>\\n2. ... <Place（放置）><8s>",
      "目标次数": {max_target_times},
      "数采负责人": "",
      "机器及环境参数": "机器人配置摘要",
      "任务级别": "简易|中等|复杂",
      "任务步骤数量": 4
    }}
  ]
}}
""".strip()
    return system, user


def extract_json_payload(text: str) -> dict[str, Any]:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, flags=re.S)
        if not match:
            raise
        return json.loads(match.group(0))


def call_qwen_json(
    system: str,
    user: str,
    api_key: str,
    model: str,
    endpoint: str,
    timeout: int = QWEN_GENERATION_TIMEOUT_SECONDS,
) -> dict[str, Any]:
    payload = {
        "model": model or "qwen3-max",
        "input": {
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ]
        },
        "parameters": {
            "result_format": "message",
            "temperature": 0.35,
            "response_format": {"type": "json_object"},
        },
    }
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = request.Request(
        endpoint,
        data=data,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with request.urlopen(req, timeout=timeout) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except HTTPError as exc:
        detail = ""
        try:
            detail = exc.read().decode("utf-8", errors="replace").strip()
        except Exception:
            detail = ""
        if detail:
            try:
                detail_payload = json.loads(detail)
            except json.JSONDecodeError:
                pass
            else:
                if isinstance(detail_payload, dict):
                    detail = "；".join(
                        str(part)
                        for part in [
                            detail_payload.get("code"),
                            detail_payload.get("message") or detail_payload.get("error"),
                        ]
                        if part
                    ) or detail
            raise ValueError(f"Qwen HTTP {exc.code}: {truncate_text(detail, 800)}") from exc
        raise ValueError(f"Qwen HTTP {exc.code}: {exc.reason}") from exc
    except URLError as exc:
        raise ValueError(f"Qwen 网络请求失败: {exc.reason}") from exc
    content = ""
    try:
        content = body["output"]["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError):
        content = body.get("output", {}).get("text") or body.get("output", "")
    if isinstance(content, list):
        content = "".join(str(part.get("text", part)) if isinstance(part, dict) else str(part) for part in content)
    return extract_json_payload(str(content))


def call_qwen(
    robots: list[RobotProfile],
    ideas: list[str],
    task_count: int,
    task_phase: str,
    api_key: str,
    model: str,
    endpoint: str,
    batch_index: int = 1,
    workspace_state: WorkbookState | None = None,
) -> list[dict[str, Any]]:
    system, user = qwen_prompt(robots, ideas, task_count, task_phase, batch_index, workspace_state=workspace_state)
    parsed = call_qwen_json(system, user, api_key, model, endpoint, timeout=QWEN_GENERATION_TIMEOUT_SECONDS)
    tasks = parsed.get("tasks", parsed if isinstance(parsed, list) else [])
    if not isinstance(tasks, list):
        raise ValueError("Qwen response JSON did not contain a tasks array")
    return [task for task in tasks if isinstance(task, dict)]


def action_key_from_label(label: str) -> str | None:
    normalized = label.strip()
    if re.fullmatch(r"\d+(?:\.\d+)?\s*s", normalized, flags=re.IGNORECASE):
        return None
    first = re.split(r"[（(]", normalized, maxsplit=1)[0].strip()
    for key, canonical in CANONICAL_ACTIONS.items():
        if first.lower() == key.lower() or normalized == canonical:
            return key
    return first or None


def extract_action_keys(steps: str) -> list[str]:
    keys: list[str] = []
    for label in re.findall(r"<([^<>]+)>", str(steps or "")):
        key = action_key_from_label(label)
        if key:
            keys.append(key)
    return keys


def match_robot(row: dict[str, Any], robots: list[RobotProfile]) -> RobotProfile | None:
    device = str(row.get("采集设备") or "").strip()
    if not robots:
        return None
    for robot in robots:
        if device == robot.name:
            return robot
    for robot in robots:
        if device and (device in robot.name or robot.name in device):
            return robot
    return robots[0] if len(robots) == 1 else None


def normalize_level(value: Any) -> str:
    text = str(value or "").strip()
    return {"易": "简易", "中": "中等", "难": "复杂"}.get(text, text if text in KNOWN_LEVELS else "简易")


def normalize_task(row: dict[str, Any], robot: RobotProfile | None, serial: int, task_phase: str) -> dict[str, Any]:
    normalized = {header: row.get(header, "") for header in TASK_HEADERS}
    phase = TASK_PHASES.get(task_phase, TASK_PHASES["pretrain"])
    max_target_times = int(phase["maxTargetTimes"])
    normalized["自动编号"] = ""
    normalized["任务ID"] = ""
    normalized["采集时长（小时）"] = ""
    normalized["任务名称"] = normalize_task_name_phase(normalized.get("任务名称"), task_phase)
    normalized["任务简述"] = strip_phase_from_brief(normalized.get("任务简述"))
    if not normalized["提交时间"]:
        normalized["提交时间"] = now_text()
    if not normalized["提交人"]:
        normalized["提交人"] = "AI需求生成器"
    if not normalized["填写日期"]:
        normalized["填写日期"] = today_midnight_text()
    if robot and not normalized["采集设备"]:
        normalized["采集设备"] = robot.name
    if robot and not normalized["机器及环境参数"]:
        normalized["机器及环境参数"] = robot.summary()

    steps = str(normalized.get("任务步骤描述") or "")
    step_count = line_count_from_steps(steps)
    normalized["任务步骤数量"] = step_count or normalized.get("任务步骤数量") or 0
    actions = extract_action_keys(steps)
    normalized["任务级别"] = normalize_level(normalized.get("任务级别") or task_level(step_count, actions))
    if normalized.get("场景域分类") not in KNOWN_CATEGORIES:
        normalized["场景域分类"] = infer_category(str(normalized.get("任务名称") or normalized.get("任务简述") or ""))
    if robot and normalized.get("采集模式") not in ["双臂", "单臂_左", "单臂_右"]:
        normalized["采集模式"] = choose_mode(robot, str(normalized.get("任务名称") or ""))
    try:
        raw_target = int(float(normalized.get("目标次数") or 0))
    except (TypeError, ValueError):
        raw_target = 0
    target = raw_target
    target_warning = ""
    if target < 1:
        target = target_for_level(str(normalized.get("任务级别")), max_target_times)
        target_warning = f"目标次数缺失或无效，已按{phase['label']}默认值修正为 {target}"
    elif target > max_target_times:
        target = max_target_times
        target_warning = f"目标次数已按{phase['label']}单任务上限修正：{raw_target} -> {target}"
    normalized["目标次数"] = target
    if target_warning:
        normalized["_target_warning"] = target_warning
    return normalized


def validate_task(row: dict[str, Any], robots: list[RobotProfile], serial: int, task_phase: str) -> dict[str, Any]:
    robot = match_robot(row, robots)
    normalized = normalize_task(row, robot, serial, task_phase)
    errors: list[str] = []
    warnings: list[str] = []
    target_warning = str(normalized.pop("_target_warning", "") or "")
    if target_warning:
        warnings.append(target_warning)

    if robot is None:
        errors.append("采集设备无法匹配输入的机器人配置")
    elif not robot.has_manipulator:
        errors.append("机器人缺少可操作末端执行器，不能生成操作类任务")

    required = ["任务名称", "任务简述", "采集设备", "采集模式", "场景域分类", "任务步骤描述"]
    for field in required:
        if not str(normalized.get(field) or "").strip():
            errors.append(f"缺少字段：{field}")

    steps = str(normalized.get("任务步骤描述") or "")
    action_keys = extract_action_keys(steps)
    unknown_actions = sorted({key for key in action_keys if key not in CANONICAL_ACTIONS})
    if unknown_actions:
        errors.append(f"包含未知动作：{', '.join(unknown_actions)}")
    if not action_keys:
        errors.append("任务步骤缺少动作标签，例如 <Grasp（抓取）><8s>")
    if seconds_from_steps(steps) <= 0:
        errors.append("任务步骤缺少秒数标签，例如 <8s>")

    declared_step_count = int(normalized.get("任务步骤数量") or 0)
    actual_step_count = line_count_from_steps(steps)
    if declared_step_count != actual_step_count and actual_step_count > 0:
        warnings.append(f"任务步骤数量已按实际行数修正：{declared_step_count} -> {actual_step_count}")
        normalized["任务步骤数量"] = actual_step_count

    if robot:
        mode = normalized.get("采集模式")
        if mode == "双臂" and not robot.is_dual_arm:
            errors.append("采集模式为双臂，但机器人配置不是双臂")
        if mode == "单臂_左" and robot.is_right_only:
            errors.append("采集模式为左臂，但机器人配置仅右臂")
        if mode == "单臂_右" and robot.is_left_only:
            errors.append("采集模式为右臂，但机器人配置仅左臂")

        action_set = set(action_keys)
        if action_set & MOBILE_ACTIONS and not robot.mobile:
            errors.append("任务包含移动/导航/搬运动作，但机器人未配置移动能力")
        if action_set & BIMANUAL_ACTIONS and not robot.is_dual_arm:
            errors.append("任务包含双臂协作类动作，但机器人不是双臂配置")
        if action_set & WHOLE_BODY_ACTIONS and not robot.whole_body:
            errors.append("任务包含蹲下/伸展动作，但机器人未配置全身能力")
        if "吸盘" in robot.end_effector and action_set & SUCTION_FORBIDDEN_ACTIONS:
            errors.append("吸盘末端不适合该任务中的布料/液体/精细或强夹持动作")
        if not robot.is_dexterous and action_set & FINE_ACTIONS:
            errors.append("非灵巧手配置不适合拧螺丝、拉链、插拔等精细动作")

        text = f"{normalized.get('任务名称')} {normalized.get('任务简述')} {steps}"
        if not robot.mobile and re.search(r"机器人移动到|导航到|巡检|跨房间|货架往返|搬运到", text):
            errors.append("文本描述需要移动底盘，但机器人未配置移动能力")
        if not robot.whole_body and re.search(r"蹲|地面|脚边|低柜|高柜|弯腰", text):
            errors.append("文本描述需要全身/低位/高位能力，但机器人未配置全身能力")
        if "吸盘" in robot.end_effector and re.search(r"毛巾|衣服|袜|布料|倒水|液体|螺丝|拉链|插头", text):
            errors.append("吸盘配置不适合文本中的柔性、液体或精细对象")

    status = "accepted" if not errors else "rejected"
    return {
        "status": status,
        "row": normalized,
        "errors": errors,
        "warnings": warnings,
        "robot": robot.name if robot else "",
    }


def normalize_duplicate_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    return "".join(char for char in text if char.isalnum())


def normalize_duplicate_task_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    match = re.match(r"^(?:预训练|后训练|pretrain|posttrain)\s*[-_：:、]?\s*(.+)$", text, flags=re.I)
    if match:
        return normalize_duplicate_text(match.group(1))
    match = re.match(r"^[预后]\s*[-_：:、]\s*(.+)$", text)
    if match:
        return normalize_duplicate_text(match.group(1))
    return normalize_duplicate_text(text)


def duplicate_requirement_keys(row: dict[str, Any]) -> set[tuple[str, ...]]:
    name = normalize_duplicate_task_name(row.get("任务名称"))
    device = normalize_duplicate_text(row.get("采集设备"))
    brief = normalize_duplicate_text(row.get("任务简述"))
    steps = normalize_duplicate_text(row.get("任务步骤描述"))
    keys: set[tuple[str, ...]] = set()
    if name:
        keys.add(("name", name))
    if name and device:
        keys.add(("name_device", device, name))
    if name and steps:
        keys.add(("name_steps", name, steps))
    if device and steps:
        keys.add(("device_steps", device, steps))
    if device and brief and steps:
        keys.add(("content_device", device, brief, steps))
    return keys


def duplicate_requirement_label(row: dict[str, Any]) -> str:
    name = str(row.get("任务名称") or "").strip() or "未命名需求"
    device = str(row.get("采集设备") or "").strip()
    return f"{name} / {device}" if device else name


def build_duplicate_requirement_index(rows: list[dict[str, Any]] | None) -> dict[tuple[str, ...], str]:
    index: dict[tuple[str, ...], str] = {}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        label = duplicate_requirement_label(row)
        for key in duplicate_requirement_keys(row):
            index.setdefault(key, label)
    return index


def duplicate_errors_for_row(
    row: dict[str, Any],
    existing_index: dict[tuple[str, ...], str],
    batch_index: dict[tuple[str, ...], str],
) -> tuple[list[str], set[tuple[str, ...]]]:
    keys = duplicate_requirement_keys(row)
    errors: list[str] = []
    existing_matches = sorted({existing_index[key] for key in keys if key in existing_index})
    if existing_matches:
        errors.append(f"与存量需求重复：{'；'.join(existing_matches[:3])}")
    batch_matches = sorted({batch_index[key] for key in keys if key in batch_index})
    if batch_matches:
        errors.append(f"与本次已生成需求重复：{'；'.join(batch_matches[:3])}")
    return errors, keys


def validate_tasks(
    rows: list[dict[str, Any]],
    robots: list[RobotProfile],
    task_phase: str,
    existing_requirements: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    existing_index = build_duplicate_requirement_index(existing_requirements)
    batch_index: dict[tuple[str, ...], str] = {}
    validations: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        item = validate_task(row, robots, index, task_phase)
        duplicate_errors, duplicate_keys = duplicate_errors_for_row(item["row"], existing_index, batch_index)
        for error in duplicate_errors:
            if error not in item["errors"]:
                item["errors"].append(error)
        if duplicate_errors:
            item["status"] = "rejected"
        label = duplicate_requirement_label(item["row"])
        for key in duplicate_keys:
            batch_index.setdefault(key, label)
        validations.append(item)
    return validations


def write_xlsx(validations: list[dict[str, Any]], robots: list[RobotProfile], output_dir: Path | None = None) -> Path:
    if OPENPYXL_IMPORT_ERROR:
        raise RuntimeError(f"openpyxl unavailable: {OPENPYXL_IMPORT_ERROR}")
    target_dir = Path(output_dir or OUTPUT_DIR)
    target_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "生成结果"
    log_ws = wb.create_sheet("校验日志")
    robot_ws = wb.create_sheet("机器人配置")

    accepted = [item["row"] for item in validations if item["status"] == "accepted"]
    ws.append(TASK_HEADERS)
    for row in accepted:
        ws.append([row.get(header, "") for header in TASK_HEADERS])

    log_ws.append(["状态", "任务名称", "采集设备", "问题", "提示"])
    for item in validations:
        row = item["row"]
        log_ws.append(
            [
                item["status"],
                row.get("任务名称", ""),
                row.get("采集设备", ""),
                "\n".join(item["errors"]),
                "\n".join(item["warnings"]),
            ]
        )

    robot_ws.append(["机器人", "末端执行器", "采集模式", "移动能力", "全身能力", "能力摘要"])
    for robot in robots:
        robot_ws.append(
            [
                robot.name,
                robot.end_effector,
                robot.arms,
                "是" if robot.mobile else "否",
                "是" if robot.whole_body else "否",
                robot.summary(),
            ]
        )

    for sheet in [ws, log_ws, robot_ws]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="283044")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            width = min(max(max_len + 2, 10), 42)
            sheet.column_dimensions[column[0].column_letter].width = width
        for row in sheet.iter_rows():
            for cell in row:
                current = cell.alignment
                cell.alignment = Alignment(
                    horizontal=current.horizontal,
                    vertical="top",
                    text_rotation=current.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=current.shrink_to_fit,
                    indent=current.indent,
                )

    filename = f"generated_data_requirements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output = target_dir / filename
    wb.save(output)
    return output


def brainstorm_ideas_response(
    body: dict[str, Any],
    current_user: dict[str, Any] | None = None,
    workspace_manager: UserWorkspaceManager | None = None,
) -> dict[str, Any]:
    robots = parse_robots(body.get("robots"))
    if not robots:
        raise ValueError("请先选择或新增至少一台机器人，再自动脑洞 idea")
    task_phase = parse_task_phase(body.get("taskPhase"))
    phase = TASK_PHASES[task_phase]
    max_target_times = int(phase["maxTargetTimes"])
    generation_count = int(float(body.get("generationTaskCount") or body.get("targetTaskCount") or 12))
    if generation_count < 1:
        raise ValueError("本次输出需求条数必须大于 0")
    if generation_count > MAX_GENERATED_TASKS_PER_REQUEST:
        raise ValueError(f"本次最多生成 {MAX_GENERATED_TASKS_PER_REQUEST} 条需求；这和单任务目标次数不是同一个概念")
    idea_count = int(float(body.get("ideaCount") or generation_count or (18 if task_phase == "pretrain" else 36)))
    idea_count = max(1, min(idea_count, MAX_GENERATED_TASKS_PER_REQUEST))
    api_key, model, endpoint, _ = resolve_qwen_request_config(body, current_user, workspace_manager)
    workspace_state = workspace_state_for(current_user, workspace_manager)
    workbook_summary = workspace_state.summary if workspace_state else WORKBOOK_SUMMARY
    rag_documents = workspace_state.rag_documents if workspace_state else RAG_DOCUMENTS
    robot_text = "\n".join(
        f"- {robot.summary()}\n  允许动作：{', '.join(derive_capabilities(robot)['allowedActions'])}"
        for robot in robots
    )
    workbook_context = workbook_context_for_prompt(example_limit=8, workbook_summary=workbook_summary, rag_documents=rag_documents)
    rag_context = build_rag_context(
        robots,
        [
            phase["label"],
            " ".join(robot.name for robot in robots),
            "通用抓取放置 家居家政 商超药店 餐饮服务 工业制造",
        ],
        rag_documents=rag_documents,
        limit_per_idea=5,
        total_limit=16,
    )
    system = (
        "你是机器人数据采集需求的任务 idea 策划器。你只输出 JSON。"
        "必须结合存量数据分布和机器人真实能力提出新 idea；不确定的能力不要假设。"
    )
    user = f"""
请自动脑洞一批新的机器人数据采集 task idea，供后续生成数据需求表使用。

任务阶段：{phase["label"]}
阶段策略：{phase["style"]}
计划生成数据需求条数：{generation_count}
单条任务目标次数上限：{max_target_times} 次
需要输出 idea 数量：{idea_count}

机器人配置：
{robot_text}

存量数据摘要：
{json.dumps(workbook_context, ensure_ascii=False, indent=2)}

存量数据 RAG 检索上下文（用于发散 idea，参考既有任务的场景、对象、动作标签和任务颗粒度，但不要照抄任务名）：
{json.dumps(rag_context, ensure_ascii=False, indent=2)}

要求：
1. 每个 idea 是短句，不写完整步骤，不写编号。
2. idea 必须符合至少一台机器人的实际能力；固定工位不要提出跨房间/巡检/货架往返 idea。
3. 预训练 idea 偏底层技能覆盖和对象泛化；后训练 idea 偏复杂约束、多对象、多场景指令。
4. 不要提出与示例任务同名或内容近似重复的 idea；若场景相近，必须更换对象、约束、动作组合或评价目标。
5. 覆盖不同场景域，优先补足存量数据中相对少的场景。

输出 JSON schema：
{{
  "ideas": ["idea 1", "idea 2"],
  "rationale": "一句话说明这些 idea 如何参考了存量数据"
}}
""".strip()
    try:
        parsed = call_qwen_json(system, user, api_key, model, endpoint, timeout=QWEN_IDEA_TIMEOUT_SECONDS)
    except TimeoutError as exc:
        raise ValueError(
            f"Qwen 自动脑洞超过 {QWEN_IDEA_TIMEOUT_SECONDS} 秒仍未返回。"
            "可以减少 idea 数量，或临时切到 qwen-turbo / qwen-plus 后重试。"
        ) from exc
    ideas = parsed.get("ideas", [])
    if not isinstance(ideas, list):
        raise ValueError("Qwen idea 响应缺少 ideas 数组")
    cleaned = []
    seen = set()
    for idea in ideas:
        text = re.sub(r"^\d+[.、]\s*", "", str(idea or "").strip())
        if text and text not in seen:
            cleaned.append(text[:80])
            seen.add(text)
    if not cleaned:
        raise ValueError("Qwen 未返回有效 idea")
    return {
        "ok": True,
        "taskPhase": task_phase,
        "phaseLabel": phase["label"],
        "model": model,
        "ideas": cleaned[:idea_count],
        "rationale": str(parsed.get("rationale") or ""),
    }


def generation_response(
    body: dict[str, Any],
    current_user: dict[str, Any] | None = None,
    workspace_manager: UserWorkspaceManager | None = None,
) -> dict[str, Any]:
    robots = parse_robots(body.get("robots"))
    if not robots:
        raise ValueError("至少需要输入一台机器人配置")

    ideas = split_task_ideas(body.get("taskIdeas"))
    task_phase = parse_task_phase(body.get("taskPhase"))
    phase = TASK_PHASES[task_phase]
    max_target_times = int(phase["maxTargetTimes"])
    fallback_count = int(body.get("countPerRobot") or 3) * max(len(robots), 1)
    match_idea_count = parse_bool(body.get("matchIdeaCount", True))
    if match_idea_count and ideas:
        generation_task_count = len(ideas)
    else:
        generation_task_count = int(float(body.get("generationTaskCount") or body.get("targetTaskCount") or fallback_count))
    if generation_task_count < 1:
        raise ValueError("本次输出需求条数必须大于 0")
    if generation_task_count > MAX_GENERATED_TASKS_PER_REQUEST:
        raise ValueError(f"本次最多生成 {MAX_GENERATED_TASKS_PER_REQUEST} 条需求；这和单任务目标次数不是同一个概念")
    api_key, model, endpoint, _ = resolve_qwen_request_config(body, current_user, workspace_manager)
    source = "qwen"
    notices: list[str] = []
    workspace_state = workspace_state_for(current_user, workspace_manager)

    generated_rows: list[dict[str, Any]] = []
    try:
        remaining = generation_task_count
        batch_index = 1
        while remaining > 0:
            batch_count = min(remaining, MAX_QWEN_TASKS_PER_BATCH)
            generated_rows.extend(
                call_qwen(
                    robots,
                    ideas,
                    batch_count,
                    task_phase,
                    api_key,
                    model,
                    endpoint,
                    batch_index,
                    workspace_state=workspace_state,
                )
            )
            remaining -= batch_count
            batch_index += 1
    except (HTTPError, URLError, TimeoutError, ValueError, json.JSONDecodeError) as exc:
        raise ValueError(f"Qwen 调用失败，未生成数据需求：{exc}") from exc

    existing_requirements = workspace_state.rag_documents if workspace_state else None
    validations = validate_tasks(generated_rows, robots, task_phase, existing_requirements=existing_requirements)
    accepted_count = sum(1 for item in validations if item["status"] == "accepted")
    rejected_count = len(validations) - accepted_count

    return {
        "ok": True,
        "source": source,
        "model": model,
        "endpoint": endpoint,
        "notices": notices,
        "summary": {
            "generated": len(validations),
            "accepted": accepted_count,
            "rejected": rejected_count,
            "requested": generation_task_count,
            "taskPhase": phase["label"],
            "maxTargetTimes": max_target_times,
        },
        "capabilities": [derive_capabilities(robot) for robot in robots],
        "items": validations,
        "rows": [item["row"] for item in validations],
    }


def export_response(
    body: dict[str, Any],
    current_user: dict[str, Any] | None = None,
    workspace_manager: UserWorkspaceManager | None = None,
) -> dict[str, Any]:
    robots = parse_robots(body.get("robots"))
    if not robots:
        raise ValueError("至少需要输入一台机器人配置")
    rows = body.get("rows")
    if not isinstance(rows, list) or not rows:
        raise ValueError("没有可导出的需求，请先生成需求")
    task_phase = parse_task_phase(body.get("taskPhase"))
    workspace_state = workspace_state_for(current_user, workspace_manager)
    existing_requirements = workspace_state.rag_documents if workspace_state else None
    validations = validate_tasks(
        [row for row in rows if isinstance(row, dict)],
        robots,
        task_phase,
        existing_requirements=existing_requirements,
    )
    if not validations:
        raise ValueError("没有可导出的有效需求行")
    accepted_count = sum(1 for item in validations if item["status"] == "accepted")
    rejected_count = len(validations) - accepted_count
    if current_user is None and workspace_manager is None:
        output_path = write_xlsx(validations, robots)
    else:
        manager = workspace_manager or WORKSPACE_MANAGER
        output_path = write_xlsx(validations, robots, manager.output_dir(current_user or LOCAL_USER))
    return {
        "ok": True,
        "downloadUrl": f"/download/{output_path.name}",
        "downloadName": output_path.name,
        "summary": {
            "generated": len(validations),
            "accepted": accepted_count,
            "rejected": rejected_count,
            "taskPhase": TASK_PHASES[task_phase]["label"],
            "maxTargetTimes": int(TASK_PHASES[task_phase]["maxTargetTimes"]),
        },
        "items": validations,
    }


class Handler(BaseHTTPRequestHandler):
    server_version = "RobotDemandPrototype/0.1"

    def log_message(self, fmt: str, *args: Any) -> None:
        print(f"[{now_text()}] {self.address_string()} {fmt % args}")

    def send_json(self, payload: dict[str, Any], status: int = 200, headers: dict[str, str] | None = None) -> None:
        data = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        for name, value in (headers or {}).items():
            self.send_header(name, value)
        self.end_headers()
        self.wfile.write(data)

    def send_file(self, path: Path, content_type: str) -> None:
        if not path.exists() or not path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def read_json_body(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length") or 0)
        raw = self.rfile.read(length) if length else b"{}"
        if len(raw) > 2_000_000:
            raise ValueError("请求体过大")
        return json.loads(raw.decode("utf-8"))

    def read_binary_body(self, limit: int = 30_000_000) -> bytes:
        length = int(self.headers.get("Content-Length") or 0)
        if length > limit:
            raise ValueError("上传文件过大")
        return self.rfile.read(length) if length else b""

    def session_token(self) -> str:
        return parse_session_cookie(self.headers.get("Cookie"))

    def current_user(self) -> dict[str, Any] | None:
        return user_from_session_token(self.session_token())

    def require_user(self) -> dict[str, Any] | None:
        user = self.current_user()
        if user:
            return user
        self.send_json({"ok": False, "error": "请先登录"}, status=401)
        return None

    def require_admin_user(self, user: dict[str, Any]) -> bool:
        try:
            AUTH_STORE.require_admin(user)
        except PermissionError as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=403)
            return False
        return True

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Cookie")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            if path == "/api/session":
                user = self.current_user()
                self.send_json(
                    {
                        "ok": True,
                        "authenticated": bool(user),
                        "user": public_session_user(user, WORKSPACE_MANAGER),
                        "defaultAdminUsername": DEFAULT_ADMIN_USERNAME,
                    }
                )
                return

            if path == "/":
                self.send_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
                return

            if path.startswith("/download/"):
                user = self.require_user()
                if not user:
                    return
                try:
                    download_path = WORKSPACE_MANAGER.resolve_download(user, Path(path).name)
                except PermissionError:
                    self.send_json({"ok": False, "error": "文件不存在或无权限"}, status=404)
                    return
                self.send_file(download_path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                return

            if path.startswith("/api/"):
                user = self.require_user()
                if not user:
                    return
                if path == "/api/health":
                    state = WORKSPACE_MANAGER.get_workspace(user)
                    qwen_config = qwen_config_response(user, WORKSPACE_MANAGER)
                    self.send_json(
                        {
                            "ok": OPENPYXL_IMPORT_ERROR is None and bool(state.summary.get("ok")),
                            "time": now_text(),
                            "user": public_session_user(user, WORKSPACE_MANAGER),
                            "sourceWorkbook": str(state.source),
                            "defaultSourceWorkbook": str(DEFAULT_SOURCE_XLSX),
                            "openpyxlError": str(OPENPYXL_IMPORT_ERROR) if OPENPYXL_IMPORT_ERROR else "",
                            "qwenConfigured": bool(qwen_config.get("configured")),
                            "qwenConfigSource": qwen_config.get("source"),
                            "qwenApiKeyMask": qwen_config.get("apiKeyMask"),
                            "qwenModel": qwen_config.get("model") or DEFAULT_QWEN_MODEL,
                            "qwenModelOptions": QWEN_MODEL_OPTIONS,
                            "qwenEndpoint": qwen_config.get("endpoint") or DEFAULT_QWEN_ENDPOINT,
                            "qwenConfig": qwen_config,
                        }
                    )
                    return
                if path == "/api/qwen/config":
                    self.send_json(qwen_config_response(user, WORKSPACE_MANAGER))
                    return
                if path == "/api/profile":
                    self.send_json(WORKSPACE_MANAGER.user_profile(user))
                    return
                if path == "/api/schema":
                    self.send_json(WORKSPACE_MANAGER.get_workspace(user).summary)
                    return
                if path == "/api/admin/users":
                    if not self.require_admin_user(user):
                        return
                    self.send_json({"ok": True, "users": AUTH_STORE.list_users(user)})
                    return
                self.send_json({"ok": False, "error": "接口不存在"}, status=404)
                return

            requested = (STATIC_DIR / path.lstrip("/")).resolve()
            if STATIC_DIR in requested.parents:
                suffix = requested.suffix.lower()
                content_type = {
                    ".css": "text/css; charset=utf-8",
                    ".js": "application/javascript; charset=utf-8",
                    ".png": "image/png",
                    ".jpg": "image/jpeg",
                    ".jpeg": "image/jpeg",
                }.get(suffix, "application/octet-stream")
                self.send_file(requested, content_type)
                return
            self.send_error(HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)
        except Exception as exc:
            self.send_json({"ok": False, "error": f"{type(exc).__name__}: {exc}"}, status=500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/auth/login":
                body = self.read_json_body()
                user = AUTH_STORE.authenticate(str(body.get("username") or ""), str(body.get("password") or ""))
                token = create_session(user)
                self.send_json(
                    {"ok": True, "authenticated": True, "user": public_session_user(user, WORKSPACE_MANAGER)},
                    headers={"Set-Cookie": session_cookie_header(token)},
                )
                return
            if parsed.path == "/api/auth/logout":
                token = self.session_token()
                if token:
                    destroy_session(token)
                self.send_json({"ok": True, "authenticated": False}, headers={"Set-Cookie": clear_session_cookie_header()})
                return

            user = self.require_user()
            if not user:
                return

            if parsed.path == "/api/admin/users":
                if not self.require_admin_user(user):
                    return
                body = self.read_json_body()
                created = AUTH_STORE.create_user(
                    user,
                    str(body.get("username") or ""),
                    str(body.get("password") or ""),
                    str(body.get("role") or "user"),
                )
                self.send_json({"ok": True, "user": created, "users": AUTH_STORE.list_users(user)})
                return
            if parsed.path == "/api/admin/users/disabled":
                if not self.require_admin_user(user):
                    return
                body = self.read_json_body()
                updated = AUTH_STORE.set_user_disabled(user, str(body.get("userId") or body.get("id") or ""), parse_bool(body.get("disabled")))
                self.send_json({"ok": True, "user": updated, "users": AUTH_STORE.list_users(user)})
                return
            if parsed.path == "/api/admin/users/password":
                if not self.require_admin_user(user):
                    return
                body = self.read_json_body()
                updated = AUTH_STORE.set_password(user, str(body.get("userId") or body.get("id") or ""), str(body.get("password") or ""))
                self.send_json({"ok": True, "user": updated})
                return
            if parsed.path == "/api/capabilities":
                body = self.read_json_body()
                robots = parse_robots(body.get("robots"))
                self.send_json({"ok": True, "capabilities": [derive_capabilities(robot) for robot in robots]})
                return
            if parsed.path == "/api/qwen/config":
                body = self.read_json_body()
                self.send_json(save_qwen_config_response(body, current_user=user, workspace_manager=WORKSPACE_MANAGER))
                return
            if parsed.path == "/api/qwen/test":
                body = self.read_json_body()
                self.send_json(qwen_config_test_response(body, current_user=user, workspace_manager=WORKSPACE_MANAGER))
                return
            if parsed.path == "/api/profile":
                body = self.read_json_body()
                profile = WORKSPACE_MANAGER.save_user_profile(user, avatar=str(body.get("avatar") or ""))
                self.send_json({"ok": True, **profile, "user": public_session_user(user, WORKSPACE_MANAGER)})
                return
            if parsed.path == "/api/generate":
                body = self.read_json_body()
                self.send_json(generation_response(body, current_user=user, workspace_manager=WORKSPACE_MANAGER))
                return
            if parsed.path == "/api/export":
                body = self.read_json_body()
                self.send_json(export_response(body, current_user=user, workspace_manager=WORKSPACE_MANAGER))
                return
            if parsed.path == "/api/ideas/brainstorm":
                body = self.read_json_body()
                self.send_json(brainstorm_ideas_response(body, current_user=user, workspace_manager=WORKSPACE_MANAGER))
                return
            if parsed.path == "/api/workbook/upload":
                raw = self.read_binary_body()
                filename, file_data = extract_multipart_file(raw, self.headers.get("Content-Type", ""))
                if not file_data:
                    raise ValueError("上传文件为空")
                saved_path = WORKSPACE_MANAGER.save_upload(user, filename, file_data)
                summary = WORKSPACE_MANAGER.set_active_workbook(user, saved_path)
                self.send_json({"ok": True, "summary": summary})
                return
            self.send_json({"ok": False, "error": "接口不存在"}, status=404)
        except json.JSONDecodeError:
            self.send_json({"ok": False, "error": "请求 JSON 无法解析"}, status=400)
        except PermissionError as exc:
            status = 401 if parsed.path == "/api/auth/login" else 403
            self.send_json({"ok": False, "error": str(exc)}, status=status)
        except ValueError as exc:
            self.send_json({"ok": False, "error": str(exc)}, status=400)
        except Exception as exc:
            self.send_json({"ok": False, "error": f"{type(exc).__name__}: {exc}"}, status=500)


def main() -> None:
    if OPENPYXL_IMPORT_ERROR:
        raise SystemExit(f"openpyxl is required: {OPENPYXL_IMPORT_ERROR}")
    port = int(os.getenv("PORT", "8787"))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"Robot demand prototype running at http://127.0.0.1:{port}/")
    print(f"Source workbook: {DEFAULT_SOURCE_XLSX}")
    print(f"Qwen model: {DEFAULT_QWEN_MODEL}")
    print(f"Qwen API: {'configured from environment' if os.getenv('DASHSCOPE_API_KEY', '').strip() else 'not configured yet'}")
    print(f"Admin username: {DEFAULT_ADMIN_USERNAME}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("Shutting down")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
